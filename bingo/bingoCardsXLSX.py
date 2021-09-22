import openpyxl, pyodbc, os
from dotenv import load_dotenv

load_dotenv(os.path.dirname(os.path.dirname(__file__)) + '\\.env')

conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + os.getenv('DEV_SERVER') + ';DATABASE=' + os.getenv('DEV_DATABASE') + ';UID=' + os.getenv('DEV_USER') + ';PWD=' + os.getenv('DEV_PASS'))

cardQuery = conn.cursor()
cardQuery.execute("""SELECT DISTINCT BingoCard FROM [dbo].[Bingo]""")

cards = cardQuery.fetchall()

for card in cards:
    bingo = conn.cursor()
    bingo.execute("""SELECT [BingoIndex]
        ,[BingoCard]
        ,[BingoNumber]
        ,[BingoPosition]
        ,[BingoCalled]
        ,[BingoMissed]
        ,[BingoDate]
    FROM [DataWarehouse].[dbo].[Bingo]
    WHERE BingoCard = """ + str(card[0]))

    tiles = bingo.fetchall()

    wb = openpyxl.load_workbook(r"C:\Users\jeremyshank\Documents\Bingo\Bingo Template.xlsx")
    ws = wb['Bingo']

    for i in range(0, 25):
        # print(tiles[i][2])
        if i < 5:
            ws.cell(column=2, row=i+3, value=tiles[i][2])
        elif i < 10 and i > 4:
            ws.cell(column=3, row=(i%5)+3, value=tiles[i][2])
        elif i < 12 and i > 9:
            ws.cell(column=4, row=(i%5)+3, value=tiles[i][2])
        elif i == 12:
            ws.cell(column=4, row=5, value="FREE")
        elif i < 15 and i > 12:
            ws.cell(column=4, row=(i%5)+3, value=tiles[i][2])
        elif i < 20 and i > 14:
            ws.cell(column=5, row=(i%5)+3, value=tiles[i][2])
        else:
            ws.cell(column=6, row=(i%5)+3, value=tiles[i][2])

    wb.save("C:\\Users\\jeremyshank\\Documents\\Bingo\\Cards\\" + str(card[0]) + ".xlsx")

    wb.close()
