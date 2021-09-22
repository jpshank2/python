import datetime
from openpyxl.styles import Font, PatternFill

x = datetime.datetime.now()
c_year = x.strftime("%Y")
c_month = x.strftime("%m")
n_month = str(int(c_month) + 1)
if int(n_month) < 10:
    n_month = "0" + n_month

def member(conn, name, wb):
    ws = wb.create_sheet("Member")
    rowNum = 2
    col = 1
    cursor = conn.cursor()
    cursor.execute("""SELECT		 
            e.ClientCode AS [Client Code]
			,E.ClientName AS [Client Name]
			,Own.StaffName AS [Client Originator]
			,Ptnr.StaffName AS [Member]
			,Mgr.StaffName AS [Client Manager]
            ,Jb.StaffName AS [Job Manager]
			,S.StaffName AS [Current Staff]
			,j.Job_Name AS [Client Project]
			,j.Job_Code AS [Job Code]
			,TP.StatusDescription AS [Job Status] 
			,TP.WorkStatusDesc AS [Workflow Status]
			,SV.ServTitle AS [Job Class]
			,T.CurrentDueDate AS [Due Date]
			,YEAR(J.Job_Period_End) AS [Job Year]
            ,E.ClientOffice AS [Client Office]
            ,J.Job_Office AS [Job Office]
			,PS.StaffName AS [Preparer]
			,RS.StaffName AS [Reviewer]
			,SV.ServTitle AS [Services]

	FROM tblEngagement AS E
		
		INNER JOIN tblJob_Header  AS J	ON j.ContIndex = e.ContIndex 
		INNER JOIN tblPortfolio_Job AS TP ON J.Job_Idx = TP.Job_Idx
		LEFT OUTER JOIN tblJob_Work_Status AS ws ON ws.StatusIndex = j.Job_WorkStatus
		LEFT OUTER JOIN tblJob_Roles P ON P.Job_Idx = J.Job_Idx And P.RoleIndex = 2 
		LEFT OUTER JOIN tblJob_Roles R ON R.Job_Idx = J.Job_Idx And R.RoleIndex = 3
		LEFT OUTER JOIN tblStaff PS ON PS.StaffIndex = P.StaffIndex
		LEFT OUTER JOIN tblStaff RS ON RS.StaffIndex = R.StaffIndex
		INNER Join tblStaff AS Ptnr ON E.ClientPartner = Ptnr.StaffIndex
		INNER join tblStaff AS Mgr ON E.ClientManager = Mgr.StaffIndex
        INNER join tblStaff AS Jb ON j.Job_Manager = Jb.StaffIndex
		INNER JOIN tblStaff AS S ON j.Job_CurrentStaff = S.StaffIndex
		INNER join tblClientOrigination AS CO ON E.ContIndex = CO.ContIndex
		INNER join tblStaff AS OWN ON CO.StaffIndex = OWN.StaffIndex
		INNER JOIN tblJob_Serv JS ON JS.Job_Idx = J.Job_Idx 
		INNER JOIN tblServices SV ON SV.ServIndex = JS.ServIndex 
		LEFT OUTER JOIN tblJob_TaxReturn T ON T.Job_Idx = J.Job_Idx
		LEFT OUTER JOIN tblJob_TaxReturn_Jurisdictions zTRJ ON zTRJ.Job_Idx = J.Job_Idx 


	WHERE e.ClientStatus IN ('ACTIVE', 'NEW') 
	AND E.ClientOffice <> 'AO' 
	AND T.CurrentDueDate >= '""" + c_year + """-""" + c_month + """-01' AND T.CurrentDueDate <= '""" + c_year + """-04-15'
	AND Ptnr.StaffName = '""" + name + """'
	ORDER BY T.CurrentDueDate""")
    rows = cursor.fetchall()
    columns = [column[0] for column in cursor.description]
    for column in columns:
        current_cell = ws.cell(row=1, column=col)
        current_cell.value = column
        current_cell.font = Font(b=True, color="FFFFFF", size=20)
        current_cell.fill = PatternFill(start_color="003D4C", end_color="003D4C", fill_type="solid")
        col += 1
    for row in rows:
        col = 1
        for entry in row:
            current_cell = ws.cell(row=rowNum, column=col)
            current_cell.value = entry
            current_cell.font = Font(color="000000", size=14)
            if rowNum % 2 == 0:
                current_cell.fill = PatternFill(start_color="BEE6B3", end_color="BEE6B3", fill_type="solid")
            col += 1
        rowNum += 1
    