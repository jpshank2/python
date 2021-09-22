import openpyxl, os

path = r"C:\Users\jeremyshank\desktop\GL for Dalton"

with os.scandir(path) as dirs:
    for entry in dirs:
        if os.path.isfile(os.path.join(path, entry)):
            print(os.path.join(path, entry))
            wb = openpyxl.load_workbook(os.path.join(path, entry), data_only=True)
            ws = wb['Report Page']

            print(type(ws['A749'].value))

            # row = 12
            # accountName = str()
            # accountNumber = str()

            # def addAccountInfo(accountNumber, accountName, row):
            #     if type(ws.cell(row=row, column=3).value) == int:
            #         ws.cell(row=row, column=1).value = accountName
            #         ws.cell(row=row, column=2).value = accountNumber

            # while ws.cell(row=row, column=3).value != 'Page 1 of 1':
            #     if ws.cell(row=row, column=3).value == 'Period':
            #         accountName = ws.cell(row=(row-3), column=4).value
            #         accountNumber = ws.cell(row=(row-3), column=3).value
            #     else:
            #         # print(accountName)
            #         # print(accountNumber)
            #         addAccountInfo(accountName, accountNumber, row)
            #     row += 1

            row = 700

            while ws.cell(row=row, column=3).value != 'Page 1 of 1':
                if ws.cell(row=row, column=1).value == None:
                    print(ws.cell(row=row, column=1).value)
                    ws.delete_rows(row, amount=1)
                    print('deleted row ' + str(row))
                row += 1
                # print(row)

            wb.save(os.path.join(r"C:\Users\jeremyshank\desktop\GL for Dalton\Cleaned GL", entry))
            wb.close()


