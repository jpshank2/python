from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
import win32com.client as wc
import os, pyodbc

load_dotenv(os.path.dirname(os.path.dirname(__file__)) + '\\.env')

devopsConn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + os.getenv('DEV_SERVER') + ';DATABASE=' + os.getenv('DEV_DATABASE') + ';UID=' + os.getenv('DEV_USER') + ';PWD=' + os.getenv('DEV_PASS'))
bmss = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};SERVER=' + os.getenv('DB_SERVER') + ';DATABASE=' + os.getenv('DB_DATABASE') + ';UID=' + os.getenv('DB_USER') + ';PWD=' + os.getenv('DB_PASS') + ';Authentication=ActiveDirectoryPassword')

hrLeaders = pd.read_sql("SELECT C.*, S.StaffEMail FROM dbo.tblCategory C INNER JOIN dbo.tblStaff S ON S.StaffName = C.CatName WHERE CatType = 'SUBDEPT' AND CatName <> 'Unknown'", bmss)

outlook = wc.Dispatch('outlook.application')

def getHRMembers(leader):
    hrMembers = pd.read_sql("""SELECT S.StaffIndex, S.StaffName 
                            FROM dbo.tblStaff S 
                            INNER JOIN dbo.tblStaffEX SE ON SE.StaffIndex = S.StaffIndex
                            WHERE S.StaffEnded IS NULL AND S.StaffType = 1 AND SE.StaffSubDepartment = '""" + leader + "'", bmss)
    
    return hrMembers

for x in range(0, 3):#hrLeaders.shape[0]):
    myHomeroom = getHRMembers(hrLeaders.iloc[x]['Category'])
    wb = Workbook()

    summaryWS = wb.create_sheet('Summary')
    summaryWS.append(['Homeroom Member', 'ROLOs Received This Quarter', 'Kudos Received This Quarter'])
    summaryCounter = 1

    roloWB = wb.create_sheet('ROLOs')
    wb.remove(wb['Sheet'])
    roloWB.append(["Recipient", "Sender", "Project", "Date", "Rating", "Retain", "Lose"])
    rolosCounter = 1

    kudosWS = wb.create_sheet('KUDOs')
    kudosWS.append(['Recipient', 'Sender', 'Cornerstone', 'Date', 'Details'])
    kudosCounter = 1

    for y in range(0, myHomeroom.shape[0]):
        summary = pd.read_sql("""SELECT DISTINCT S.SubRecipient, COALESCE(ROLOSCount, 0) AS [ROLOSCount], COALESCE(KUDOSCount, 0) AS [KUDOSCount]
FROM MandM.Submissions S
	LEFT JOIN (SELECT SubRecipient, COUNT(*) AS [ROLOSCount] FROM MandM.Submissions WHERE SubType IN (1,2) AND SubRecipient = """ + str(myHomeroom.iloc[y]['StaffIndex']) + """ AND DATEDIFF(QUARTER, SubDate, GETDATE()) = 0 GROUP BY SubRecipient) R ON R.SubRecipient = S.SubRecipient
	LEFT JOIN (SELECT SubRecipient, COUNT(*) AS [KUDOSCount] FROM MandM.Submissions WHERE SubType = 5 AND SubRecipient = """ + str(myHomeroom.iloc[y]['StaffIndex']) + """ AND DATEDIFF(QUARTER, SubDate, GETDATE()) = 0 GROUP BY SubRecipient) K ON K.SubRecipient = S.SubRecipient
WHERE S.SubRecipient = """ + str(myHomeroom.iloc[y]['StaffIndex']), devopsConn)

        rolos = pd.read_sql("""SELECT SubSender
                                    ,SubRecipient
                                    ,CASE
                                        WHEN SubRating = 3 THEN 'Thumbs Up'
                                        WHEN SubRating = 2 THEN 'Okay'
                                        ELSE 'Thumbs Down'
                                    END AS Rating
                                    ,SubHeading AS [Project]
                                    ,SubNotes1 AS [Retain]
                                    ,SubNotes2 AS [Lose]
                                    ,CAST(SubDate AS DATE) AS SubDate
                                FROM MandM.Submissions
                                WHERE SubType IN (1, 2)
                                AND DATEDIFF(WEEK, SubDate, GETDATE()) < 4 AND SubRecipient = """ + str(myHomeroom.iloc[y]['StaffIndex']), devopsConn)
        
        kudos = pd.read_sql("""SELECT SubSender
                                    ,SubRecipient
                                    ,SubHeading AS [Cornerstone]
                                    ,SubNotes1 AS [Details]
                                    ,CAST(SubDate AS DATE) AS SubDate
                                FROM MandM.Submissions
                                WHERE SubType = 5
                                AND DATEDIFF(WEEK, SubDate, GETDATE()) < 4 AND SubRecipient = """ + str(myHomeroom.iloc[y]['StaffIndex']), devopsConn)
        
        if rolos.shape[0] == 0:
            roloWB.append([myHomeroom.iloc[y]['StaffName'], 'No ROLOs Received', '', '', '', '', ''])
            rolosCounter += 1
        else:
            for z in range(0, rolos.shape[0]):
                roloWB.append([myHomeroom.iloc[y]['StaffName'], pd.read_sql("SELECT StaffName FROM dbo.tblStaff WHERE StaffIndex = " + str(rolos.iloc[z]['SubSender']), bmss).iloc[0]['StaffName'], rolos.iloc[z]['Project'], rolos.iloc[z]['SubDate'], rolos.iloc[z]['Rating'], rolos.iloc[z]['Retain'], rolos.iloc[z]['Lose']])
                rolosCounter += 1
        
        if kudos.shape[0] == 0:
            kudosWS.append([myHomeroom.iloc[y]['StaffName'], 'No KUDOs Received', '', '', ''])
            kudosCounter += 1
        else:
            for alpha in range(0, kudos.shape[0]):
                kudosWS.append([myHomeroom.iloc[y]['StaffName'], pd.read_sql("SELECT StaffName FROM dbo.tblStaff WHERE StaffIndex = " + str(kudos.iloc[alpha]['SubSender']), bmss).iloc[0]['StaffName'], kudos.iloc[alpha]['Cornerstone'], kudos.iloc[alpha]['SubDate'], kudos.iloc[alpha]['Details']])
                kudosCounter += 1
        
        for beta in range(0, summary.shape[0]):
            summaryWS.append([myHomeroom.iloc[y]['StaffName'], summary.iloc[beta]['ROLOSCount'], summary.iloc[beta]['KUDOSCount']])
            summaryCounter += 1

    style = TableStyleInfo(name='TableStyleMedium11', showRowStripes=True, showColumnStripes=False)

    roloTab = Table(displayName="ROLOs", ref="A1:G" + str(rolosCounter))
    roloTab.tableStyleInfo = style
    roloWB.add_table(roloTab)

    kudosTab = Table(displayName="KUDOs", ref="A1:E" + str(kudosCounter))
    kudosTab.tableStyleInfo = style
    kudosWS.add_table(kudosTab)

    summaryTab = Table(displayName='Summary', ref="A1:C" + str(summaryCounter))
    summaryTab.tableStyleInfo = style
    summaryWS.add_table(summaryTab)

    wb.save(r'C:\users\jeremyshank\desktop\\' + hrLeaders.iloc[x]['CatName'] + '.xlsx')
    wb.close()
